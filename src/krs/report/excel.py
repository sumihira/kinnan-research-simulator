from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from krs.simulation.experiment import ExperimentResult
from krs.simulation.runner import GoldfishRunResult


SUMMARY_SHEET_NAME: Final[str] = "Summary"
GAMES_SHEET_NAME: Final[str] = "Games"

HEADER_FILL_COLOR: Final[str] = "1F4E78"
HEADER_FONT_COLOR: Final[str] = "FFFFFF"
SECTION_FILL_COLOR: Final[str] = "D9EAF7"
LABEL_FILL_COLOR: Final[str] = "F3F6F9"
WIN_FILL_COLOR: Final[str] = "E2F0D9"
NON_WIN_FILL_COLOR: Final[str] = "FFFFFF"
BORDER_COLOR: Final[str] = "D9E1F2"

SUMMARY_TITLE_ROW: Final[int] = 1
SUMMARY_START_ROW: Final[int] = 3
GAMES_HEADER_ROW: Final[int] = 1
GAMES_START_ROW: Final[int] = 2


@dataclass(frozen=True, slots=True)
class ExcelExperimentReporter:
    """
    Creates a formatted Excel workbook from ExperimentResult.

    The reporter serializes existing simulation configuration, summary, and
    game results. It does not recalculate statistics or modify the supplied
    ExperimentResult.
    """

    title: str = "Kinnan Research Simulator Report"

    def __post_init__(self) -> None:
        if not self.title.strip():
            raise ValueError("title must not be empty.")

    def create_workbook(
        self,
        result: ExperimentResult,
    ) -> Workbook:
        """
        Create an in-memory Excel workbook for one experiment.
        """
        workbook = Workbook()

        summary_sheet = workbook.active

        if summary_sheet is None:
            raise RuntimeError(
                "Workbook did not create an active worksheet."
            )

        summary_sheet.title = SUMMARY_SHEET_NAME

        games_sheet = workbook.create_sheet(
            title=GAMES_SHEET_NAME,
        )

        self._populate_summary_sheet(
            summary_sheet,
            result,
        )
        self._populate_games_sheet(
            games_sheet,
            result,
        )

        workbook.active = 0

        return workbook

    def write(
        self,
        result: ExperimentResult,
        path: str | Path,
    ) -> Path:
        """
        Create and save an Excel report.

        Missing parent directories are created automatically.
        """
        output_path = Path(path)

        if output_path.exists() and output_path.is_dir():
            raise ValueError(
                f"Excel report path is a directory: {output_path}"
            )

        if output_path.suffix.casefold() != ".xlsx":
            raise ValueError(
                "Excel report path must use the .xlsx extension."
            )

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook = self.create_workbook(result)

        try:
            workbook.save(output_path)
        finally:
            workbook.close()

        return output_path

    def _populate_summary_sheet(
        self,
        sheet: Worksheet,
        result: ExperimentResult,
    ) -> None:
        """
        Populate and format the Summary worksheet.
        """
        self._write_summary_title(sheet)

        current_row = SUMMARY_START_ROW

        current_row = self._write_section_header(
            sheet,
            row=current_row,
            label="Configuration",
        )

        for label, value in self._configuration_rows(result):
            self._write_summary_value_row(
                sheet,
                row=current_row,
                label=label,
                value=value,
            )
            current_row += 1

        current_row = self._write_section_header(
            sheet,
            row=current_row,
            label="Summary",
        )

        for label, value in self._summary_rows(result):
            self._write_summary_value_row(
                sheet,
                row=current_row,
                label=label,
                value=value,
            )
            current_row += 1

        self._apply_summary_number_formats(sheet)
        self._configure_summary_sheet(
            sheet,
            final_row=current_row - 1,
        )

    def _populate_games_sheet(
        self,
        sheet: Worksheet,
        result: ExperimentResult,
    ) -> None:
        """
        Populate and format the Games worksheet.
        """
        headers = self._game_headers()

        for column, header in enumerate(
            headers,
            start=1,
        ):
            cell = sheet.cell(
                row=GAMES_HEADER_ROW,
                column=column,
                value=header,
            )
            self._style_header_cell(cell)

        for game_id, game_result in enumerate(
            result.game_results,
        ):
            self._write_game_row(
                sheet,
                row=game_id + GAMES_START_ROW,
                game_id=game_id,
                result=game_result,
            )

        self._configure_games_sheet(
            sheet,
            game_count=len(result.game_results),
            column_count=len(headers),
        )

    def _write_summary_title(
        self,
        sheet: Worksheet,
    ) -> None:
        """
        Write the merged workbook title.
        """
        sheet.merge_cells(
            start_row=SUMMARY_TITLE_ROW,
            start_column=1,
            end_row=SUMMARY_TITLE_ROW,
            end_column=2,
        )

        title_cell = sheet.cell(
            row=SUMMARY_TITLE_ROW,
            column=1,
            value=self.title,
        )
        title_cell.font = Font(
            bold=True,
            size=16,
            color=HEADER_FONT_COLOR,
        )
        title_cell.fill = PatternFill(
            fill_type="solid",
            fgColor=HEADER_FILL_COLOR,
        )
        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        title_cell.border = self._thin_border()

        sheet.cell(
            row=SUMMARY_TITLE_ROW,
            column=2,
        ).border = self._thin_border()

        sheet.row_dimensions[SUMMARY_TITLE_ROW].height = 28

    @classmethod
    def _write_section_header(
        cls,
        sheet: Worksheet,
        *,
        row: int,
        label: str,
    ) -> int:
        """
        Write a merged section header.

        Returns the next writable row.
        """
        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=2,
        )

        cell = sheet.cell(
            row=row,
            column=1,
            value=label,
        )
        cell.font = Font(
            bold=True,
        )
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=SECTION_FILL_COLOR,
        )
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )
        cell.border = cls._thin_border()

        sheet.cell(
            row=row,
            column=2,
        ).border = cls._thin_border()

        return row + 1

    @classmethod
    def _write_summary_value_row(
        cls,
        sheet: Worksheet,
        *,
        row: int,
        label: str,
        value: object,
    ) -> None:
        """
        Write one Summary label-value row.
        """
        label_cell = sheet.cell(
            row=row,
            column=1,
            value=label,
        )
        value_cell = sheet.cell(
            row=row,
            column=2,
            value=value,
        )

        label_cell.font = Font(
            bold=True,
        )
        label_cell.fill = PatternFill(
            fill_type="solid",
            fgColor=LABEL_FILL_COLOR,
        )

        for cell in (
            label_cell,
            value_cell,
        ):
            cell.border = cls._thin_border()
            cell.alignment = Alignment(
                vertical="center",
            )

    @classmethod
    def _write_game_row(
        cls,
        sheet: Worksheet,
        *,
        row: int,
        game_id: int,
        result: GoldfishRunResult,
    ) -> None:
        """
        Write one Goldfish game result.
        """
        values = (
            game_id,
            result.turns_started,
            result.kinnan_activations,
            result.reached_turn_limit,
            result.game_over,
            result.winner,
        )

        is_win = (
            result.game_over
            and result.winner is not None
        )

        fill_color = (
            WIN_FILL_COLOR
            if is_win
            else NON_WIN_FILL_COLOR
        )

        for column, value in enumerate(
            values,
            start=1,
        ):
            cell = sheet.cell(
                row=row,
                column=column,
                value=value,
            )
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=fill_color,
            )
            cell.border = cls._thin_border()
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    @staticmethod
    def _configuration_rows(
        result: ExperimentResult,
    ) -> tuple[tuple[str, object], ...]:
        """
        Return configuration values for the Summary worksheet.
        """
        config = result.config

        return (
            (
                "Strategy",
                config.strategy_name,
            ),
            (
                "Games requested",
                config.games,
            ),
            (
                "Maximum turns",
                config.max_turns,
            ),
            (
                "Seed",
                config.seed,
            ),
            (
                "Workers",
                config.workers,
            ),
            (
                "Mulligan enabled",
                config.mulligan_enabled,
            ),
            (
                "Save replays",
                config.save_replays,
            ),
        )

    @staticmethod
    def _summary_rows(
        result: ExperimentResult,
    ) -> tuple[tuple[str, object], ...]:
        """
        Return aggregate values for the Summary worksheet.
        """
        summary = result.summary

        return (
            (
                "Games completed",
                summary.games_completed,
            ),
            (
                "Wins",
                summary.wins,
            ),
            (
                "Non-wins",
                summary.non_wins,
            ),
            (
                "Win rate",
                summary.win_rate,
            ),
            (
                "Turn-limit games",
                summary.turn_limit_games,
            ),
            (
                "Total turns started",
                summary.total_turns_started,
            ),
            (
                "Average turns started",
                summary.average_turns_started,
            ),
            (
                "Total Kinnan activations",
                summary.total_kinnan_activations,
            ),
            (
                "Average Kinnan activations",
                summary.average_kinnan_activations,
            ),
            (
                "Fastest win turn",
                summary.fastest_win_turn,
            ),
        )

    @staticmethod
    def _game_headers() -> tuple[str, ...]:
        """
        Return the Games worksheet header order.
        """
        return (
            "game_id",
            "turns_started",
            "kinnan_activations",
            "reached_turn_limit",
            "game_over",
            "winner",
        )

    @classmethod
    def _apply_summary_number_formats(
        cls,
        sheet: Worksheet,
    ) -> None:
        """
        Apply number formats to percentage and average values.
        """
        win_rate_row = cls._find_label_row(
            sheet,
            "Win rate",
        )
        average_turns_row = cls._find_label_row(
            sheet,
            "Average turns started",
        )
        average_activations_row = cls._find_label_row(
            sheet,
            "Average Kinnan activations",
        )

        sheet.cell(
            row=win_rate_row,
            column=2,
        ).number_format = "0.00%"

        sheet.cell(
            row=average_turns_row,
            column=2,
        ).number_format = "0.000"

        sheet.cell(
            row=average_activations_row,
            column=2,
        ).number_format = "0.000"

    @staticmethod
    def _configure_summary_sheet(
        sheet: Worksheet,
        *,
        final_row: int,
    ) -> None:
        """
        Apply worksheet-level Summary configuration.
        """
        sheet.column_dimensions["A"].width = 31
        sheet.column_dimensions["B"].width = 22

        sheet.freeze_panes = "A3"
        sheet.sheet_view.showGridLines = False

        if final_row >= SUMMARY_START_ROW:
            sheet.auto_filter.ref = (
                f"A{SUMMARY_START_ROW}:B{final_row}"
            )

    @classmethod
    def _configure_games_sheet(
        cls,
        sheet: Worksheet,
        *,
        game_count: int,
        column_count: int,
    ) -> None:
        """
        Apply worksheet-level Games configuration.
        """
        last_row = game_count + GAMES_HEADER_ROW

        if game_count > 0:
            table = Table(
                displayName="GamesTable",
                ref=f"A1:F{last_row}",
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        else:
            sheet.auto_filter.ref = "A1:F1"

        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False

        widths = {
            "A": 12,
            "B": 16,
            "C": 21,
            "D": 21,
            "E": 14,
            "F": 24,
        }

        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        if game_count == 0:
            return

        for row in sheet.iter_rows(
            min_row=GAMES_START_ROW,
            max_row=last_row,
            min_col=1,
            max_col=column_count,
        ):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

    @classmethod
    def _style_header_cell(
        cls,
        cell: Cell,
    ) -> None:
        """
        Apply shared header styling.
        """
        cell.font = Font(
            bold=True,
            color=HEADER_FONT_COLOR,
        )
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=HEADER_FILL_COLOR,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        cell.border = cls._thin_border()

    @staticmethod
    def _find_label_row(
        sheet: Worksheet,
        label: str,
    ) -> int:
        """
        Return the row containing one Summary label.
        """
        for row in range(
            1,
            sheet.max_row + 1,
        ):
            if sheet.cell(
                row=row,
                column=1,
            ).value == label:
                return row

        raise ValueError(
            f"Summary label not found: {label}"
        )

    @staticmethod
    def _thin_border() -> Border:
        """
        Return the shared thin cell border.
        """
        side = Side(
            style="thin",
            color=BORDER_COLOR,
        )

        return Border(
            left=side,
            right=side,
            top=side,
            bottom=side,
        )