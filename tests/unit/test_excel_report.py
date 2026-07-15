from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from krs.report.excel import ExcelExperimentReporter
from krs.simulation.experiment import (
    ExperimentResult,
    SimulationSummary,
)
from krs.simulation.runner import GoldfishRunResult
from krs.simulation.simulation_config import SimulationConfig


def create_game_result(
    *,
    turns_started: int,
    kinnan_activations: int,
    reached_turn_limit: bool = False,
    game_over: bool = False,
    winner: str | None = None,
) -> GoldfishRunResult:
    return GoldfishRunResult(
        turns_started=turns_started,
        kinnan_activations=kinnan_activations,
        reached_turn_limit=reached_turn_limit,
        game_over=game_over,
        winner=winner,
    )


def create_experiment_result() -> ExperimentResult:
    config = SimulationConfig(
        strategy_name="combo",
        games=3,
        max_turns=8,
        seed=12345,
        mulligan_enabled=False,
        save_replays=True,
        workers=4,
    )

    game_results = (
        create_game_result(
            turns_started=3,
            kinnan_activations=2,
            game_over=True,
            winner="Player",
        ),
        create_game_result(
            turns_started=5,
            kinnan_activations=1,
            game_over=True,
            winner="プレイヤー",
        ),
        create_game_result(
            turns_started=8,
            kinnan_activations=0,
            reached_turn_limit=True,
        ),
    )

    summary = SimulationSummary.from_results(
        games_requested=config.games,
        results=game_results,
    )

    return ExperimentResult(
        config=config,
        game_results=game_results,
        summary=summary,
    )


def create_no_win_experiment_result() -> ExperimentResult:
    config = SimulationConfig(
        games=1,
        seed=None,
    )

    game_results = (
        create_game_result(
            turns_started=6,
            kinnan_activations=0,
            reached_turn_limit=True,
        ),
    )

    summary = SimulationSummary.from_results(
        games_requested=config.games,
        results=game_results,
    )

    return ExperimentResult(
        config=config,
        game_results=game_results,
        summary=summary,
    )


def find_summary_row(
    sheet: Worksheet,
    label: str,
) -> int:
    for row in range(
        1,
        sheet.max_row + 1,
    ):
        if sheet.cell(
            row=row,
            column=1,
        ).value == label:
            return row

    raise AssertionError(
        f"Summary label not found: {label}"
    )


def find_summary_value(
    sheet: Worksheet,
    label: str,
) -> object:
    row = find_summary_row(
        sheet,
        label,
    )

    return sheet.cell(
        row=row,
        column=2,
    ).value


def test_create_workbook_creates_expected_sheets() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        assert workbook.sheetnames == [
            "Summary",
            "Games",
        ]
    finally:
        workbook.close()


def test_create_workbook_activates_summary_sheet() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        assert workbook.active is not None
        assert workbook.active.title == "Summary"
    finally:
        workbook.close()


def test_summary_sheet_contains_custom_title() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter(
        title="KRS Monte Carlo Analysis",
    ).create_workbook(result)

    try:
        sheet = workbook["Summary"]

        assert sheet["A1"].value == "KRS Monte Carlo Analysis"
        assert "A1:B1" in {
            str(cell_range)
            for cell_range in sheet.merged_cells.ranges
        }
    finally:
        workbook.close()


def test_summary_title_is_formatted() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        title_cell = workbook["Summary"]["A1"]

        assert title_cell.font.bold is True
        assert title_cell.font.size == 16
        assert title_cell.alignment.horizontal == "center"
        assert title_cell.alignment.vertical == "center"
        assert title_cell.fill.fill_type == "solid"
    finally:
        workbook.close()


def test_summary_sheet_contains_section_headers() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Summary"]

        assert sheet["A3"].value == "Configuration"

        summary_row = find_summary_row(
            sheet,
            "Summary",
        )

        assert summary_row > 3
        assert (
            sheet.cell(
                row=summary_row,
                column=1,
            ).font.bold
            is True
        )
    finally:
        workbook.close()


def test_summary_sheet_contains_configuration_values() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Summary"]

        assert find_summary_value(
            sheet,
            "Strategy",
        ) == "combo"
        assert find_summary_value(
            sheet,
            "Games requested",
        ) == 3
        assert find_summary_value(
            sheet,
            "Maximum turns",
        ) == 8
        assert find_summary_value(
            sheet,
            "Seed",
        ) == 12345
        assert find_summary_value(
            sheet,
            "Workers",
        ) == 4
        assert find_summary_value(
            sheet,
            "Mulligan enabled",
        ) is False
        assert find_summary_value(
            sheet,
            "Save replays",
        ) is True
    finally:
        workbook.close()


def test_summary_sheet_contains_summary_values() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Summary"]

        assert find_summary_value(
            sheet,
            "Games completed",
        ) == 3
        assert find_summary_value(
            sheet,
            "Wins",
        ) == 2
        assert find_summary_value(
            sheet,
            "Non-wins",
        ) == 1
        assert find_summary_value(
            sheet,
            "Win rate",
        ) == pytest.approx(2 / 3)
        assert find_summary_value(
            sheet,
            "Turn-limit games",
        ) == 1
        assert find_summary_value(
            sheet,
            "Total turns started",
        ) == 16
        assert find_summary_value(
            sheet,
            "Average turns started",
        ) == pytest.approx(16 / 3)
        assert find_summary_value(
            sheet,
            "Total Kinnan activations",
        ) == 3
        assert find_summary_value(
            sheet,
            "Average Kinnan activations",
        ) == pytest.approx(1.0)
        assert find_summary_value(
            sheet,
            "Fastest win turn",
        ) == 3
    finally:
        workbook.close()


def test_summary_sheet_supports_none_values() -> None:
    result = create_no_win_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Summary"]

        assert find_summary_value(
            sheet,
            "Seed",
        ) is None
        assert find_summary_value(
            sheet,
            "Fastest win turn",
        ) is None
    finally:
        workbook.close()


def test_summary_sheet_uses_expected_number_formats() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Summary"]

        win_rate_row = find_summary_row(
            sheet,
            "Win rate",
        )
        average_turns_row = find_summary_row(
            sheet,
            "Average turns started",
        )
        average_activations_row = find_summary_row(
            sheet,
            "Average Kinnan activations",
        )

        assert sheet.cell(
            row=win_rate_row,
            column=2,
        ).number_format == "0.00%"

        assert sheet.cell(
            row=average_turns_row,
            column=2,
        ).number_format == "0.000"

        assert sheet.cell(
            row=average_activations_row,
            column=2,
        ).number_format == "0.000"
    finally:
        workbook.close()


def test_summary_sheet_is_configured() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Summary"]

        assert sheet.freeze_panes == "A3"
        assert sheet.sheet_view.showGridLines is False
        assert sheet.column_dimensions["A"].width == 31
        assert sheet.column_dimensions["B"].width == 22
        assert sheet.auto_filter.ref is not None
    finally:
        workbook.close()


def test_games_sheet_contains_expected_headers() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Games"]

        headers = tuple(
            sheet.cell(
                row=1,
                column=column,
            ).value
            for column in range(
                1,
                7,
            )
        )

        assert headers == (
            "game_id",
            "turns_started",
            "kinnan_activations",
            "reached_turn_limit",
            "game_over",
            "winner",
        )
    finally:
        workbook.close()


def test_games_headers_are_formatted() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Games"]

        for column in range(
            1,
            7,
        ):
            cell = sheet.cell(
                row=1,
                column=column,
            )

            assert cell.font.bold is True
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.vertical == "center"
            assert cell.fill.fill_type == "solid"
    finally:
        workbook.close()


def test_games_sheet_contains_ordered_game_ids() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Games"]

        game_ids = tuple(
            sheet.cell(
                row=row,
                column=1,
            ).value
            for row in range(
                2,
                5,
            )
        )

        assert game_ids == (
            0,
            1,
            2,
        )
    finally:
        workbook.close()


def test_games_sheet_contains_first_game_result() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Games"]

        row = tuple(
            sheet.cell(
                row=2,
                column=column,
            ).value
            for column in range(
                1,
                7,
            )
        )

        assert row == (
            0,
            3,
            2,
            False,
            True,
            "Player",
        )
    finally:
        workbook.close()


def test_games_sheet_contains_unicode_winner() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Games"]

        assert sheet["F3"].value == "プレイヤー"
    finally:
        workbook.close()


def test_games_sheet_contains_non_win_result() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Games"]

        row = tuple(
            sheet.cell(
                row=4,
                column=column,
            ).value
            for column in range(
                1,
                7,
            )
        )

        assert row == (
            2,
            8,
            0,
            True,
            False,
            None,
        )
    finally:
        workbook.close()


def test_games_sheet_contains_excel_table() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Games"]

        assert "GamesTable" in sheet.tables
        assert sheet.tables["GamesTable"].ref == "A1:F4"
    finally:
        workbook.close()


def test_games_sheet_is_configured() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Games"]

        assert sheet.freeze_panes == "A2"
        assert sheet.sheet_view.showGridLines is False
        assert sheet.column_dimensions["A"].width == 12
        assert sheet.column_dimensions["B"].width == 16
        assert sheet.column_dimensions["C"].width == 21
        assert sheet.column_dimensions["D"].width == 21
        assert sheet.column_dimensions["E"].width == 14
        assert sheet.column_dimensions["F"].width == 24
    finally:
        workbook.close()


def test_games_rows_are_center_aligned() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Games"]

        for row in range(
            2,
            5,
        ):
            for column in range(
                1,
                7,
            ):
                cell = sheet.cell(
                    row=row,
                    column=column,
                )

                assert cell.alignment.horizontal == "center"
                assert cell.alignment.vertical == "center"
    finally:
        workbook.close()


def test_win_and_non_win_rows_have_different_fills() -> None:
    result = create_experiment_result()

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        sheet = workbook["Games"]

        win_fill = sheet["A2"].fill.fgColor.rgb
        non_win_fill = sheet["A4"].fill.fgColor.rgb

        assert win_fill != non_win_fill
    finally:
        workbook.close()


def test_write_creates_xlsx_file(
    tmp_path: Path,
) -> None:
    result = create_experiment_result()
    output_path = tmp_path / "experiment.xlsx"

    returned_path = ExcelExperimentReporter().write(
        result,
        output_path,
    )

    assert returned_path == output_path
    assert output_path.is_file()


def test_written_workbook_can_be_loaded(
    tmp_path: Path,
) -> None:
    result = create_experiment_result()
    output_path = tmp_path / "experiment.xlsx"

    ExcelExperimentReporter().write(
        result,
        output_path,
    )

    workbook = load_workbook(output_path)

    try:
        assert workbook.sheetnames == [
            "Summary",
            "Games",
        ]
        assert workbook["Summary"]["A1"].value == (
            "Kinnan Research Simulator Report"
        )
        assert workbook["Games"]["F3"].value == "プレイヤー"
    finally:
        workbook.close()


def test_write_creates_parent_directories(
    tmp_path: Path,
) -> None:
    result = create_experiment_result()

    output_path = (
        tmp_path
        / "reports"
        / "excel"
        / "experiment.xlsx"
    )

    ExcelExperimentReporter().write(
        result,
        output_path,
    )

    assert output_path.is_file()


def test_write_accepts_uppercase_xlsx_extension(
    tmp_path: Path,
) -> None:
    result = create_experiment_result()
    output_path = tmp_path / "REPORT.XLSX"

    ExcelExperimentReporter().write(
        result,
        output_path,
    )

    assert output_path.is_file()


def test_write_rejects_directory_path(
    tmp_path: Path,
) -> None:
    result = create_experiment_result()

    with pytest.raises(
        ValueError,
        match="Excel report path is a directory",
    ):
        ExcelExperimentReporter().write(
            result,
            tmp_path,
        )


@pytest.mark.parametrize(
    "filename",
    (
        "report.xls",
        "report.csv",
        "report.html",
        "report",
    ),
)
def test_write_rejects_invalid_extension(
    tmp_path: Path,
    filename: str,
) -> None:
    result = create_experiment_result()

    with pytest.raises(
        ValueError,
        match=(
            "Excel report path must use "
            "the .xlsx extension."
        ),
    ):
        ExcelExperimentReporter().write(
            result,
            tmp_path / filename,
        )


@pytest.mark.parametrize(
    "title",
    (
        "",
        " ",
        "\t",
        "\n",
    ),
)
def test_reporter_rejects_empty_title(
    title: str,
) -> None:
    with pytest.raises(
        ValueError,
        match="title must not be empty.",
    ):
        ExcelExperimentReporter(
            title=title,
        )


def test_create_workbook_does_not_modify_experiment_result() -> None:
    result = create_experiment_result()

    original_config = result.config
    original_summary = result.summary
    original_results = result.game_results

    workbook = ExcelExperimentReporter().create_workbook(
        result,
    )

    try:
        assert result.config is original_config
        assert result.summary is original_summary
        assert result.game_results is original_results
    finally:
        workbook.close()


def test_write_does_not_modify_experiment_result(
    tmp_path: Path,
) -> None:
    result = create_experiment_result()

    original_config = result.config
    original_summary = result.summary
    original_results = result.game_results

    ExcelExperimentReporter().write(
        result,
        tmp_path / "experiment.xlsx",
    )

    assert result.config is original_config
    assert result.summary is original_summary
    assert result.game_results is original_results